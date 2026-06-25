"""Router Laporan (Modul 10): data JSON + export Excel."""

import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from ..services import reports as R

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _build(report_type: str, db: Session, start: date | None, end: date | None) -> dict:
    fn = R.REPORTS.get(report_type)
    if not fn:
        raise HTTPException(404, detail="Jenis laporan tidak dikenal")
    return fn(db, start, end)


@router.get("/{report_type}")
def get_report(
    report_type: str,
    start: date | None = Query(None),
    end: date | None = Query(None),
    db: Session = Depends(get_db),
):
    return _build(report_type, db, start, end)


@router.get("/{report_type}/export")
def export_report(
    report_type: str,
    start: date | None = Query(None),
    end: date | None = Query(None),
    db: Session = Depends(get_db),
):
    """Export laporan ke file .xlsx (openpyxl)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = _build(report_type, db, start, end)

    wb = Workbook()
    ws = wb.active
    ws.title = report_type[:30]

    # Judul
    ws.append([data["title"]])
    ws["A1"].font = Font(bold=True, size=14)
    if start or end:
        ws.append([f"Periode: {start or '-'} s/d {end or '-'}"])
    ws.append([])

    # Header kolom
    cols = data["columns"]
    header_row = ws.max_row + 1
    ws.append([c["label"] for c in cols])
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    # Baris data
    for row in data["rows"]:
        ws.append([row.get(c["key"], "") for c in cols])

    # Ringkasan
    ws.append([])
    for k, v in data["summary"].items():
        ws.append([k, v])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{report_type}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
